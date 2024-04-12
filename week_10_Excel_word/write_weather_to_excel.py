from openpyxl import Workbook

week_temperatures = {
    'Monday': 54,
    'Tuesday': 60,
    'Wednesday': 62,
    'Thursday': 57,
    'Friday': 71,
}

# writing to worksheet, date in first column, temp in second

workbook = Workbook()  # indicates empty workbook we are writing to
worksheet = workbook.active
worksheet.title = 'Daily Temperatures'  # any sort of naming conventions needs to go AFTER worksheet is set or the
# program won't know where to go/what to do with that information

worksheet.cell(1, 1, 'Day')
worksheet.cell(1, 2, 'Temperature (F)') # naming data columns

row_index = 2  # this keeps track of which row we are writing to - modified to 2 in order to add titles to first row -
# noted above - as Day and Temp

for day, temp in week_temperatures.items():  # this tells our program where to start and which one to start with,
    # and where to put the data we are indicating
    worksheet.cell(row_index, 1, day)
    worksheet.cell(row_index, 2, temp)  # row_index is set to 1
    row_index += 1  # adds 1 to row_index, bc python starts at 0

workbook.save('temperatures.xlsx')
