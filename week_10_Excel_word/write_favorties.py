from openpyxl import Workbook

favorite_foods = ['Spaghetti', 'Sandwich', 'Cabbage', 'Muffins', 'Cheese Burgers']
favorite_colors = ['Green', 'Pink', 'Black', 'Purple', 'Red']

workbook = Workbook()  # creating a new workbook
# won't exist on computer yet

worksheet = workbook.active  # writing to active Excel file, which we are building
worksheet.title = 'Favorite Things'  # titles file

worksheet.cell(1, 1, 'Favorite Foods')
worksheet.cell(1, 2, 'Favorite Colors')  # these add "titles" for our lists in rows and columns indicated

for index, food in enumerate(favorite_foods):  # writes data in our built lists to Excel file, starting with where we
    # indicate in loop
    worksheet.cell(index + 2, 1, food)

for index, color in enumerate(favorite_colors):  # writing/adding data from created list favorite colors to worksheet.
    # Indicating I want data from the list to be written in the second column, starting with the second row due to
    # title being added at beginning
    worksheet.cell(index + 2, 2, color)

workbook.save('favorites.xlsx')  # worksheet we are working on MUST be closed to update - if it is open when we run
# the program, it WILL NOT update this needs to go last - saves workbook based on code set before it. you can tell the
# file is open if there is a temporary file in your project - it looks like ~$filename.xlsx. once its closed it will
# disappear
